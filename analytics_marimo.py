import marimo

__generated_with = "0.9.0"
app = marimo.App(width="medium")


@app.cell
def _():
    import marimo as mo
    return (mo,)


@app.cell
async def _():
    # Environment detection + shared stdlib imports (defined once).
    import sys
    import io
    import traceback
    IS_WASM = ("pyodide" in sys.modules) or sys.platform == "emscripten"
    # openpyxl isn't in Pyodide's preloaded set, so install it on demand in the
    # browser. Cells that import openpyxl depend on openpyxl_ready, so this runs
    # first.
    if IS_WASM:
        import micropip
        await micropip.install(["openpyxl", "plotly"])
    openpyxl_ready = True
    return IS_WASM, io, openpyxl_ready, traceback


@app.cell
def _(mo):
    mo.md(
        """
        # Editor portfolio analytics

        Upload your **QTS report** (Excel) — the DOI column is read for you — and get
        a full picture of the papers you've handled, computed **in your browser** so
        your file never leaves your machine:

        - **scientific spread** — an interactive topic wheel (filter by subfield), plus field/subfield/topic tables
        - **performance** — citations, FWCI and Altmetric, with field-normalised percentiles, in sortable tables and a citations × FWCI × Altmetric bubble chart
        - **PubPeer** — comments and retraction / expression-of-concern flags
        - **collaboration reach**, **top-10 countries**, and **monthly trends**
        - optional **who's citing you** — journals, institutions, countries and the top citing papers

        Altmetric needs a configured proxy; PubPeer works from a normal (non-corporate)
        connection; the "who's citing you" lookup is opt-in as it uses OpenAlex credits.
        """
    )
    return


@app.cell
def _():
    # ── Config ────────────────────────────────────────────────────────────────
    BRAND = "5B2C6F"
    RECENT_DAYS = 60
    TOP_N = 12
    WHEEL_TOP_N = 32
    # Paste your deployed Cloudflare Worker URL here to add Altmetric scores.
    # Leave it empty ("") to run without Altmetric.
    WORKER_URL = "https://altmetric-proxy.leiajudge.workers.dev"
    COUNTRY_NAMES = {
        "US": "United States", "GB": "United Kingdom", "CN": "China",
        "DE": "Germany", "FR": "France", "JP": "Japan", "CA": "Canada",
        "AU": "Australia", "IT": "Italy", "ES": "Spain", "NL": "Netherlands",
        "CH": "Switzerland", "SE": "Sweden", "KR": "South Korea", "IN": "India",
        "BR": "Brazil", "BE": "Belgium", "DK": "Denmark", "AT": "Austria",
        "NO": "Norway", "FI": "Finland", "IL": "Israel", "IE": "Ireland",
        "SG": "Singapore", "PT": "Portugal", "PL": "Poland", "RU": "Russia",
        "MX": "Mexico", "ZA": "South Africa", "NZ": "New Zealand", "GR": "Greece",
        "CZ": "Czechia", "HU": "Hungary", "TR": "Turkey", "TW": "Taiwan",
        "HK": "Hong Kong", "SA": "Saudi Arabia", "AE": "United Arab Emirates",
        "AR": "Argentina", "CL": "Chile", "IR": "Iran", "TH": "Thailand",
        "MY": "Malaysia", "LU": "Luxembourg", "SI": "Slovenia", "EE": "Estonia",
    }
    return BRAND, COUNTRY_NAMES, RECENT_DAYS, TOP_N, WHEEL_TOP_N, WORKER_URL


@app.cell
def _(COUNTRY_NAMES, io, openpyxl_ready):
    import re
    from openpyxl import load_workbook

    def clean_doi(raw):
        s = str(raw).strip()
        s = re.sub(r"^https?://(dx\.)?doi\.org/", "", s, flags=re.I)
        s = re.sub(r"^doi:\s*", "", s, flags=re.I)
        return s.strip()

    def country_name(code):
        return COUNTRY_NAMES.get((code or "").upper(), code or "Unknown")

    def dois_from_xlsx_bytes(data):
        out = []
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            try:
                header = list(next(it))
            except StopIteration:
                continue
            idx = next((i for i, h in enumerate(header)
                        if isinstance(h, str) and h.strip().lower() == "doi"), None)
            rows = list(it)
            if idx is None:
                best, hits = None, 0
                for i in range(len(header)):
                    n = sum(1 for r in rows[:25] if i < len(r)
                            and isinstance(r[i], str) and re.search(r"10\.\d{4,}/", r[i]))
                    if n > hits:
                        best, hits = i, n
                idx = best
            if idx is None:
                continue
            for r in rows:
                if idx < len(r) and r[idx]:
                    out.append(clean_doi(r[idx]))
        # de-dupe, keep order, valid DOIs only
        seen, res = set(), []
        for d in out:
            k = d.lower()
            if d.lower().startswith("10.") and k not in seen:
                seen.add(k)
                res.append(d)
        return res

    def extract(doi, data):
        pt = data.get("primary_topic") or {}
        codes = set()
        inst_names = set()
        for a in data.get("authorships", []) or []:
            for c in a.get("countries", []) or []:
                if c:
                    codes.add(c.upper())
            for inst in a.get("institutions", []) or []:
                if inst.get("country_code"):
                    codes.add(inst["country_code"].upper())
                if inst.get("display_name"):
                    inst_names.add(inst["display_name"])
        cnp = data.get("citation_normalized_percentile") or {}
        pctl = cnp.get("value")
        return {
            "doi": doi,
            "oa_id": data.get("id") or "",
            "title": data.get("title") or "",
            "date": data.get("publication_date") or "",
            "citations": data.get("cited_by_count"),
            "fwci": data.get("fwci"),
            "percentile": pctl,  # 0-1; higher = cited more than that fraction of peers
            "top1": bool(cnp.get("is_in_top_1_percent")),
            "top10": bool(cnp.get("is_in_top_10_percent")),
            "domain": (pt.get("domain") or {}).get("display_name") or "Unclassified",
            "field": (pt.get("field") or {}).get("display_name") or "Unclassified",
            "subfield": (pt.get("subfield") or {}).get("display_name") or "Unclassified",
            "topic": pt.get("display_name") or "Unclassified",
            "countries": sorted(codes),
            "n_countries": len(codes),
            "inst_names": sorted(inst_names),
        }

    return clean_doi, country_name, dois_from_xlsx_bytes, extract, re


@app.cell
def _(IS_WASM, WORKER_URL):
    # ── Cross-environment OpenAlex fetch ──────────────────────────────────────
    # Browser: pyodide.http.pyfetch (async). Local: urllib (stdlib).
    # A contact email puts requests in OpenAlex's faster "polite pool".
    MAILTO = "editor-analytics@users.noreply.github.com"

    async def fetch_openalex(doi, mailto=MAILTO):
        select = ("id,title,publication_date,cited_by_count,fwci,primary_topic,"
                  "authorships,type,citation_normalized_percentile")
        url = "https://api.openalex.org/works/doi:{}?select={}&mailto={}".format(
            doi, select, mailto)
        if IS_WASM:
            from pyodide.http import pyfetch
            resp = await pyfetch(url)
            if resp.status == 404:
                return None
            if resp.status != 200:
                return None
            return await resp.json()
        else:
            import urllib.request
            import urllib.error
            import json
            try:
                with urllib.request.urlopen(url, timeout=15) as r:
                    return json.loads(r.read().decode("utf-8"))
            except urllib.error.HTTPError as e:
                if e.code == 404:
                    return None
                raise

    async def fetch_altmetric(dois):
        # Ask the Cloudflare Worker for Altmetric scores. Returns {doi_lower: score}.
        # Empty dict if WORKER_URL isn't set or anything goes wrong (non-fatal).
        import json
        if not WORKER_URL:
            return {}
        payload = json.dumps({"dois": dois})
        if IS_WASM:
            from pyodide.http import pyfetch
            resp = await pyfetch(WORKER_URL, method="POST",
                                 headers={"Content-Type": "application/json"},
                                 body=payload)
            if resp.status != 200:
                return {}
            data = await resp.json()
        else:
            import urllib.request
            req = urllib.request.Request(
                WORKER_URL, data=payload.encode("utf-8"),
                headers={"Content-Type": "application/json"}, method="POST")
            with urllib.request.urlopen(req, timeout=90) as r:
                data = json.loads(r.read().decode("utf-8"))
        return data.get("scores", {}) or {}

    async def _oa_json(url):
        if IS_WASM:
            from pyodide.http import pyfetch
            resp = await pyfetch(url)
            return (await resp.json()) if resp.status == 200 else None
        import urllib.request
        import json
        try:
            with urllib.request.urlopen(url, timeout=30) as r:
                return json.loads(r.read().decode("utf-8"))
        except Exception:
            return None

    async def fetch_citing(items, self_insts, progress=None):
        # items: list of (oa_id, cited_by_count).
        # FAST PATH: batch papers into OR'd group_by queries — 3 calls per batch
        # of BATCH papers instead of ~1 per paper (a ~25x cut in requests).
        # SAFETY NET: we already know how many citations to expect (the sum of
        # cited_by_count), so if the batched result captures far fewer than that,
        # we assume the batched filter misbehaved and automatically redo it one
        # paper at a time. Fast when batching works, correct when it doesn't.
        import asyncio
        from collections import Counter
        BATCH = 25          # stay well inside OpenAlex's OR-list limits
        MIN_COVERAGE = 0.5  # below this share of expected citations, fall back
        base = "https://api.openalex.org/works"
        GROUPS = (("primary_location.source.id", "journals"),
                  ("authorships.institutions.id", "institutions"),
                  ("authorships.countries", "countries"))
        cited = [((w or "").rsplit("/", 1)[-1], c) for w, c in items]
        cited = [(w, c) for w, c in cited if w.startswith("W") and c]
        expected = sum(c for _, c in cited)
        _dbg = {"mode": "batched", "papers": len(cited), "calls": 0, "ok": 0,
                "expected": expected, "captured": 0, "sample": None}

        async def _grouped(filter_val, gb):
            url = ("{}?filter=cites:{}&group_by={}&per-page=200&mailto={}"
                   .format(base, filter_val, gb, MAILTO))
            if _dbg["sample"] is None:
                _dbg["sample"] = url
            data = None
            for _try in range(3):
                data = await _oa_json(url)
                if data is not None:
                    break
            _dbg["calls"] += 1
            if data:
                _dbg["ok"] += 1
            return ((data or {}).get("group_by") or [])

        async def _run(batch_size, report):
            counters = {"journals": Counter(), "institutions": Counter(),
                        "countries": Counter()}
            jids = {}   # journal display name -> OpenAlex source id
            for start in range(0, len(cited), batch_size):
                chunk = cited[start:start + batch_size]
                filt = "|".join(w for w, _ in chunk)
                for gb, name in GROUPS:
                    for g in await _grouped(filt, gb):
                        key = g.get("key_display_name") or g.get("key")
                        if key and str(key).lower() not in ("unknown", "none"):
                            counters[name][key] += g.get("count", 0)
                            if name == "journals" and g.get("key"):
                                jids.setdefault(key, str(g["key"]).rsplit("/", 1)[-1])
                if report and progress:
                    for _ in chunk:
                        progress()
                await asyncio.sleep(0)
            return counters, jids

        counters, jids = await _run(BATCH, report=True)
        captured = sum(counters["journals"].values())
        _dbg["captured"] = captured

        # Safety net: batched result clearly incomplete -> redo per paper.
        if expected and captured < MIN_COVERAGE * expected:
            _dbg["mode"] = "per-paper fallback"
            _dbg["batched_captured"] = captured
            counters, jids = await _run(1, report=False)
            _dbg["captured"] = sum(counters["journals"].values())

        top_journals = counters["journals"].most_common(100)

        # Journal-level impact: OpenAlex "2-year mean citedness" — computed the
        # same way as an impact factor (the JIF itself is proprietary Clarivate
        # data and has no open API). ~2 extra calls for 100 journals.
        impact = {}
        wanted = [jids[n] for n, _ in top_journals if jids.get(n)]
        for start in range(0, len(wanted), 50):
            ids = "|".join(wanted[start:start + 50])
            url = ("https://api.openalex.org/sources?filter=ids.openalex:{}"
                   "&select=id,summary_stats&per-page=50&mailto={}".format(ids, MAILTO))
            data = None
            for _try in range(3):
                data = await _oa_json(url)
                if data is not None:
                    break
            _dbg["calls"] += 1
            for s in ((data or {}).get("results") or []):
                sid = str(s.get("id") or "").rsplit("/", 1)[-1]
                val = (s.get("summary_stats") or {}).get("2yr_mean_citedness")
                if sid and isinstance(val, (int, float)):
                    impact[sid] = round(val, 1)

        return {
            "journals": [(n, c, ("nature" in str(n).lower()),
                          impact.get(jids.get(n, ""), ""))
                         for n, c in top_journals],
            "institutions": [(n, c, (n in self_insts))
                             for n, c in counters["institutions"].most_common(100)],
            "countries": counters["countries"].most_common(100),
            "_dbg": _dbg,
        }

    async def fetch_pubpeer(dois):
        # Direct browser call to PubPeer's public endpoint. Works from a normal
        # (residential) connection; fails quietly on networks that block it.
        # Returns {doi_lower: {"comments": n, "url": ..., "flag": ...}}.
        import json
        out = {}
        endpoint = "https://pubpeer.com/v3/publications?devkey=PubMedChrome"
        for start in range(0, len(dois), 40):
            chunk = dois[start:start + 40]
            payload = json.dumps({"version": "1.6.2", "browser": "Chrome",
                                  "urls": [], "dois": chunk})
            data = None
            try:
                if IS_WASM:
                    from pyodide.http import pyfetch
                    resp = await pyfetch(
                        endpoint, method="POST",
                        headers={"Content-Type": "application/json;charset=UTF-8"},
                        body=payload)
                    if resp.status == 200:
                        data = await resp.json()
                else:
                    import urllib.request
                    req = urllib.request.Request(
                        endpoint, data=payload.encode("utf-8"),
                        headers={"Content-Type": "application/json;charset=UTF-8"},
                        method="POST")
                    with urllib.request.urlopen(req, timeout=60) as r:
                        data = json.loads(r.read().decode("utf-8"))
            except Exception:
                data = None
            for fb in ((data or {}).get("feedbacks") or []):
                doi = (fb.get("id") or "").lower()
                if not doi:
                    continue
                flag = ""
                for u in (fb.get("updates") or []):
                    t = (u.get("type") or "").upper()
                    if "RETRACT" in t:
                        flag = "Retracted"
                    elif "CONCERN" in t and not flag:
                        flag = "Expression of concern"
                out[doi] = {
                    "comments": fb.get("total_comments") or 0,
                    "url": fb.get("url") or ("https://pubpeer.com/publications/" + doi),
                    "flag": flag,
                }
        return out

    return fetch_altmetric, fetch_citing, fetch_openalex, fetch_pubpeer


@app.cell
def _(WHEEL_TOP_N, country_name, io, openpyxl_ready):
    # ── Analysis + rendering (pure, no network) ───────────────────────────────
    from collections import Counter
    import numpy as np
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import plotly.graph_objects as go
    from openpyxl import Workbook
    from openpyxl.styles import Font

    def summarise(records):
        n = len(records)
        field = Counter(r["field"] for r in records)
        subfield = Counter(r["subfield"] for r in records)
        topic = Counter(r["topic"] for r in records)
        domain = Counter(r["domain"] for r in records)
        country = Counter(c for r in records for c in r["countries"])
        intl = sum(1 for r in records if r["n_countries"] >= 2)
        dated = [r["date"] for r in records if r["date"]]
        return {
            "n": n, "field": field, "subfield": subfield, "topic": topic,
            "domain": domain, "country": country, "intl": intl,
            "range": ("{} to {}".format(min(dated), max(dated)) if dated else "-"),
        }

    def pctl_label(r):
        v = r.get("percentile")
        if not isinstance(v, (int, float)):
            return ""
        top = (1.0 - v) * 100.0
        return "Top {:.0f}%".format(top) if top >= 1 else "Top 1%"

    def perf_rows(records, key):
        """All papers as rows, sorted by the given metric (desc).
        key is one of 'citations', 'fwci', 'altmetric'."""
        def val(r):
            v = r.get(key)
            return v if isinstance(v, (int, float)) else -1
        return [{"DOI": r["doi"], "Title": r["title"], "Published": r["date"],
                 "Citations": r["citations"], "FWCI": r["fwci"],
                 "Altmetric": r.get("altmetric"), "Field rank": pctl_label(r)}
                for r in sorted(records, key=val, reverse=True)]

    def wheel_png(records, only_subfield=None):
        if only_subfield and only_subfield != "All":
            records = [r for r in records if r["subfield"] == only_subfield]
        tc = Counter(r["topic"] for r in records)
        total = len(tc)
        # Biggest topics first; each topic is its own wedge and its own colour.
        selected = [t for t, _ in tc.most_common(WHEEL_TOP_N)]
        counts = [tc[t] for t in selected]
        if not selected:
            return None
        # A distinct colour per topic: stitch together several qualitative maps.
        palette = []
        for cmap_name in ("tab20", "tab20b", "tab20c"):
            palette.extend(plt.get_cmap(cmap_name).colors)
        colors = [palette[i % len(palette)] for i in range(len(selected))]

        ang = np.linspace(0, 2 * np.pi, len(selected), endpoint=False)
        rmax = max(counts)
        fig = plt.figure(figsize=(9, 9))
        ax = fig.add_subplot(111, projection="polar")
        ax.set_theta_offset(np.pi / 2)
        ax.set_theta_direction(-1)
        ax.bar(ang, counts, width=(2 * np.pi / len(selected)) * 0.95,
               color=colors, edgecolor="white", linewidth=1.0, alpha=0.9)
        ax.set_xticks([])
        ax.set_yticklabels([])
        ax.spines["polar"].set_visible(False)
        ax.set_ylim(0, rmax * 1.10)
        for a, c in zip(ang, counts):
            if c >= max(2, rmax * 0.25):
                ax.text(a, c / 2, str(c), ha="center", va="center",
                        fontsize=6.5, color="white", weight="bold")
        # Legend of topics (truncate long names so it stays readable).
        def _short(t):
            return t if len(t) <= 42 else t[:39] + "…"
        handles = [plt.Rectangle((0, 0), 1, 1, color=colors[i])
                   for i in range(len(selected))]
        ax.legend(handles, [_short(t) for t in selected], loc="center left",
                  bbox_to_anchor=(1.02, 0.5), fontsize=13, frameon=False,
                  title="Topic", ncol=2 if len(selected) > 16 else 1)
        shown = "top {} of {}".format(len(selected), total) if total > len(selected) else "all {}".format(len(selected))
        scope = "" if (not only_subfield or only_subfield == "All") else " — {}".format(only_subfield)
        ax.set_title("Portfolio by topic ({}){}".format(shown, scope),
                     fontsize=13, pad=24, weight="bold")
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
        plt.close(fig)
        return buf.getvalue()

    def build_xlsx(records):
        s = summarise(records)
        wb = Workbook()
        ws = wb.active
        ws.title = "Science"
        for col, (name, counter) in enumerate(
                [("Field", s["field"]), ("Subfield", s["subfield"]),
                 ("Topic", s["topic"]), ("Country", s["country"])]):
            base = col * 3
            ws.cell(row=1, column=base + 1, value=name).font = Font(bold=True)
            ws.cell(row=1, column=base + 2, value="Papers").font = Font(bold=True)
            for i, (k, v) in enumerate(counter.most_common(), start=2):
                label = country_name(k) if name == "Country" else k
                ws.cell(row=i, column=base + 1, value=label)
                ws.cell(row=i, column=base + 2, value=v)
        wb.create_sheet("Papers")
        wp = wb["Papers"]
        heads = ["DOI", "Title", "Published", "Citations", "FWCI",
                 "Altmetric", "Field", "Subfield", "Topic", "Countries"]
        for c, h in enumerate(heads, 1):
            wp.cell(row=1, column=c, value=h).font = Font(bold=True)
        for i, r in enumerate(records, start=2):
            for c, key in enumerate(
                    ["doi", "title", "date", "citations", "fwci", "altmetric",
                     "field", "subfield", "topic"], 1):
                wp.cell(row=i, column=c, value=r.get(key))
            wp.cell(row=i, column=10, value=", ".join(r["countries"]))
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    def bubble_fig(records):
        # x = citations, y = FWCI, size = Altmetric (by rank, so it's dynamic).
        xs, ys, sizes, texts, links = [], [], [], [], []
        for r in records:
            xs.append(r["citations"] if isinstance(r["citations"], (int, float)) else 0)
            ys.append(r["fwci"] if isinstance(r["fwci"], (int, float)) else 0)
            a = r.get("altmetric")
            sizes.append(a if isinstance(a, (int, float)) else 0)
            texts.append((r["title"] or "")[:90])
            links.append("https://doi.org/" + r["doi"])
        n = len(sizes)
        # True area-proportional sizing (like Excel): bubble AREA scales with the
        # Altmetric score, so outliers genuinely dominate. Plotly's sizemode="area"
        # + sizeref does exactly this. sizeref is set so the biggest score maps to
        # ~MAXPX pixels across; a size floor keeps zero/low papers just visible.
        MAXPX = 60.0
        FLOOR = 6.0
        smax = max(sizes) if sizes else 0
        sizeref = (2.0 * smax / (MAXPX ** 2)) if smax > 0 else 1.0
        raw = [(s if isinstance(s, (int, float)) and s > 0 else 0.0) for s in sizes]
        cdata = [[links[i], (sizes[i] if sizes[i] else "—")] for i in range(n)]
        fig = go.Figure(go.Scatter(
            x=xs, y=ys, mode="markers",
            marker=dict(size=raw, sizemode="area", sizeref=sizeref, sizemin=FLOOR,
                        color=ys, colorscale="Viridis", showscale=False,
                        line=dict(width=0.5, color="white"), opacity=0.7),
            customdata=cdata, text=texts,
            hovertemplate=("<b>%{text}</b><br>Citations: %{x}<br>FWCI: %{y:.2f}"
                           "<br>Altmetric: %{customdata[1]}<extra></extra>")))
        fig.update_layout(xaxis_title="Citations", yaxis_title="FWCI",
                          height=520, template="simple_white", dragmode="select",
                          margin=dict(l=50, r=20, t=20, b=45))
        return fig

    def country_fig(records):
        cc = Counter(c for r in records for c in r["countries"])
        top = cc.most_common(10)[::-1]
        fig = go.Figure(go.Bar(
            x=[n for _, n in top], y=[country_name(c) for c, _ in top],
            orientation="h", marker_color="#5B2C6F",
            hovertemplate="%{y}: %{x} papers<extra></extra>"))
        fig.update_layout(height=380, template="simple_white",
                          xaxis_title="Papers", margin=dict(l=10, r=10, t=10, b=30))
        return fig

    def collab_stats(records):
        n = len(records)
        solo = sum(1 for r in records if r["n_countries"] <= 1)
        avg = (sum(r["n_countries"] for r in records) / n) if n else 0
        partners = Counter(c for r in records if r["n_countries"] >= 2
                           for c in r["countries"])
        return {"solo": solo, "intl": n - solo, "avg": avg,
                "partners": [(country_name(c), v) for c, v in partners.most_common(8)]}

    def bench_summary(records):
        withp = sum(1 for r in records if isinstance(r.get("percentile"), (int, float)))
        return {"n_pct": withp,
                "top1": sum(1 for r in records if r.get("top1")),
                "top10": sum(1 for r in records if r.get("top10"))}

    def trend_fig(records):
        import statistics
        by_month = {}
        for r in records:
            m = (r["date"] or "")[:7]  # YYYY-MM
            if len(m) == 7 and m[:4].isdigit() and m[5:7].isdigit():
                by_month.setdefault(m, []).append(r)
        months = sorted(by_month)
        counts = [len(by_month[m]) for m in months]
        med = []
        for m in months:
            fs = [r["fwci"] for r in by_month[m] if isinstance(r["fwci"], (int, float))]
            med.append(round(statistics.median(fs), 2) if fs else None)
        fig = go.Figure()
        fig.add_bar(x=months, y=counts, name="Papers", marker_color="#B7A4CE")
        fig.add_scatter(x=months, y=med, name="Median FWCI", mode="lines+markers",
                        line=dict(color="#5B2C6F", width=3), yaxis="y2")
        fig.update_layout(height=380, template="simple_white",
                          xaxis=dict(title="Month", type="category"),
                          yaxis=dict(title="Papers"),
                          yaxis2=dict(title="Median FWCI", overlaying="y",
                                      side="right", showgrid=False),
                          legend=dict(orientation="h", y=1.12),
                          margin=dict(l=40, r=45, t=30, b=40))
        return fig

    return (bench_summary, bubble_fig, build_xlsx, collab_stats, country_fig,
            perf_rows, summarise, trend_fig, wheel_png)


@app.cell
def _(mo):
    file = mo.ui.file(filetypes=[".xlsx"], label="Upload your QTS report (.xlsx)")
    run = mo.ui.run_button(label="Build my analytics")
    cite_run = mo.ui.run_button(label="Look up who's citing me")
    mo.vstack([file, run])
    return cite_run, file, run


@app.cell
async def _(dois_from_xlsx_bytes, fetch_altmetric, fetch_openalex, fetch_pubpeer, extract, file, mo, run, traceback, WORKER_URL):
    # Gate the heavy work behind the button + an uploaded file.
    import asyncio
    mo.stop(not run.value, mo.md("*Upload your QTS report, then press **Build my analytics**.*"))
    mo.stop(not file.value, mo.md("⚠️ No file uploaded yet."))

    records = []
    _err = None
    try:
        _dois = dois_from_xlsx_bytes(file.contents())
        if not _dois:
            _err = "No DOIs found in that spreadsheet."
        else:
            with mo.status.progress_bar(
                total=len(_dois), title="Fetching from OpenAlex",
                subtitle="First run downloads the Python packages — hang on a moment.",
            ) as bar:
                for _i, _d in enumerate(_dois):
                    _data = await fetch_openalex(_d)
                    if _data:
                        records.append(extract(_d, _data))
                    bar.update(subtitle="{} of {} papers".format(_i + 1, len(_dois)))
                    # Yield to the browser so the progress bar actually repaints.
                    await asyncio.sleep(0)
            # Optional Altmetric scores (only if a Worker URL is configured).
            if WORKER_URL and records:
                with mo.status.spinner(title="Fetching Altmetric scores…"):
                    try:
                        _scores = await fetch_altmetric([r["doi"] for r in records])
                    except Exception:
                        _scores = {}
                for r in records:
                    r["altmetric"] = _scores.get(r["doi"].lower())
            # PubPeer comments (direct browser call; non-fatal if blocked)
            if records:
                with mo.status.spinner(title="Checking PubPeer…"):
                    try:
                        _pp = await fetch_pubpeer([r["doi"] for r in records])
                    except Exception:
                        _pp = {}
                for r in records:
                    _info = _pp.get(r["doi"].lower()) or {}
                    r["pp_comments"] = _info.get("comments", 0)
                    r["pp_url"] = _info.get("url", "")
                    r["pp_flag"] = _info.get("flag", "")
    except Exception:
        _err = traceback.format_exc()

    (mo.md("### ⚠️ Error while fetching\n\n```\n{}\n```".format(_err)) if _err
     else mo.md("Fetched **{}** of {} papers.".format(len(records), len(_dois))))
    return (records,)


@app.cell
async def _(cite_run, fetch_citing, mo, records):
    citing = None
    if records and cite_run.value:
        _items = [(r["oa_id"], r["citations"] or 0) for r in records]
        _self_insts = set()
        for _r in records:
            for _nm in _r.get("inst_names", []):
                _self_insts.add(_nm)
        with mo.status.progress_bar(total=len(_items),
                                    title="Finding who cites your papers…") as _cbar:
            try:
                citing = await fetch_citing(_items, _self_insts,
                                            progress=lambda: _cbar.update())
            except Exception:
                citing = None
    return (citing,)


@app.cell
def _(bubble_fig, mo, records, summarise):
    # UI elements live in their own cell so their .value changes drive updates.
    mo.stop(not records, mo.md(""))
    _subs = sorted(summarise(records)["subfield"].keys())
    subfield_dd = mo.ui.dropdown(options=["All"] + _subs, value="All",
                                 label="Filter the topic wheel by subfield")
    bubble = mo.ui.plotly(bubble_fig(records))
    return bubble, subfield_dd


@app.cell
def _(bench_summary, bubble, build_xlsx, collab_stats, country_fig,
      perf_rows, mo, records, subfield_dd, summarise, traceback,
      trend_fig, wheel_png):
    mo.stop(not records, mo.md(""))
    try:
        _s = summarise(records)
        _intl_pct = (_s["intl"] / _s["n"]) if _s["n"] else 0
        _summary = mo.md(
            "### {} papers · {}\n\n"
            "**{}** fields · **{}** subfields · **{}** topics · "
            "**{}** countries · **{:.0%}** international".format(
                _s["n"], _s["range"], len(_s["field"]), len(_s["subfield"]),
                len(_s["topic"]), len(_s["country"]), _intl_pct))

        # Drag a box over bubbles -> persistent, clickable links below the chart.
        # (A single click usually won't register as a selection; dragging does.)
        _links_all = ["https://doi.org/" + r["doi"] for r in records]
        _sel_urls = []
        try:
            for _p in (bubble.value or []):
                _u = _p.get("customdata")
                if isinstance(_u, (list, tuple)):
                    _u = _u[0] if _u else None
                if not _u:
                    _idx = _p.get("pointNumber")
                    if _idx is None:
                        _idx = _p.get("pointIndex")
                    if _idx is None:
                        _idx = _p.get("point_number", _p.get("point_index"))
                    if isinstance(_idx, int) and 0 <= _idx < len(_links_all):
                        _u = _links_all[_idx]
                if _u and _u not in _sel_urls:
                    _sel_urls.append(_u)
        except Exception:
            pass
        if _sel_urls:
            _clicked = mo.md("**Selected papers ({}):**\n\n".format(len(_sel_urls))
                             + "\n".join("- [{}]({})".format(u, u) for u in _sel_urls))
        else:
            _clicked = mo.md("*Drag a box across bubbles to list them here as "
                             "clickable links — the hover box shows each paper.*")

        _wheel = mo.image(wheel_png(records, subfield_dd.value), width=680)
        _country = mo.ui.plotly(country_fig(records))

        # DOIs render as clickable links (open the paper) but stay sortable.
        _fmt = {"DOI": lambda v: mo.md("[{}](https://doi.org/{})".format(v, v))}
        _tab = lambda key: mo.ui.table(perf_rows(records, key), selection=None,
                                       pagination=True, page_size=15, format_mapping=_fmt)
        _tabs = mo.ui.tabs({"Citations": _tab("citations"),
                            "FWCI": _tab("fwci"), "Altmetric": _tab("altmetric")})

        _dl = mo.download(data=build_xlsx(records),
                          filename="portfolio_analytics.xlsx",
                          label="Download full workbook (.xlsx)")

        # PubPeer comments (papers with any comments/flags)
        _pprows = sorted(
            [{"DOI": r["doi"], "Title": r["title"],
              "Comments": r.get("pp_comments", 0), "Flag": r.get("pp_flag", ""),
              "Link": r.get("pp_url", "")}
             for r in records if r.get("pp_comments")],
            key=lambda x: x["Comments"], reverse=True)
        if _pprows:
            _pp_view = mo.ui.table(
                _pprows, selection=None, pagination=True, page_size=25,
                format_mapping={
                    "DOI": lambda v: mo.md("[{}](https://doi.org/{})".format(v, v)),
                    "Link": lambda v: mo.md("[PubPeer]({})".format(v)) if v else ""})
        else:
            _pp_view = mo.md("*No PubPeer comments found on these papers (or PubPeer "
                             "wasn't reachable from this network).*")

        # Benchmarks (field/year-normalised percentiles)
        _bs = bench_summary(records)
        _bench = mo.md("**Field-normalised standing:** **{}** in the top 1% · "
                       "**{}** in the top 10%  *(of {} papers with percentile data)*".format(
                           _bs["top1"], _bs["top10"], _bs["n_pct"]))

        # Collaboration reach
        _cs = collab_stats(records)
        _partners = ", ".join("{} ({})".format(nm, ct) for nm, ct in _cs["partners"]) or "—"
        _collab = mo.md(
            "### Collaboration reach\n\n"
            "**{}** international · **{}** single-country · **{:.1f}** countries per "
            "paper on average\n\n**Most frequent partner countries:** {}".format(
                _cs["intl"], _cs["solo"], _cs["avg"], _partners))

        # Time trend
        _trend = mo.ui.plotly(trend_fig(records))

        _out = mo.vstack([
            _summary, _bench,
            mo.md("### PubPeer comments"), _pp_view,
            mo.md("### Citations × FWCI × Altmetric"),
            mo.md("*Each bubble is a paper. Bubble area = Altmetric score, so "
                  "outliers stand out. Hover for details; **drag a box** over "
                  "bubbles to list them with links below.*"),
            bubble, _clicked,
            mo.md("### Over time"),
            mo.md("*Bars = papers per month; line = median FWCI per month.*"), _trend,
            mo.md("### Top 10 countries"), _country,
            _collab,
            mo.md("### Topic wheel"), subfield_dd, _wheel,
            mo.md("### Performance — ranked by each metric (DOIs are clickable)"), _tabs,
            _dl,
        ])
    except Exception:
        _out = mo.md("### ⚠️ Error while building the report\n\n```\n{}\n```".format(
            traceback.format_exc()))
    _out
    return


@app.cell
def _(cite_run, citing, country_name, mo, records):
    # Separate cell so the (optional) citation lookup doesn't hold up the dashboard.
    mo.stop(not records, mo.md(""))
    if not cite_run.value:
        _view = mo.vstack([
            mo.md("*Optional — this makes extra OpenAlex calls (which draw on your "
                  "daily free credits), so it's off by default. Press to run it.*"),
            cite_run])
    elif citing is None:
        _view = mo.md("*Looking up who cites your papers…*")
    elif citing.get("journals") or citing.get("institutions"):
        _jrows = [{"Journal": nm, "Citations": ct, "Self (Nature)": "✓" if slf else "",
                   "2-yr mean citedness": imp}
                  for nm, ct, slf, imp in citing["journals"]]
        _irows = [{"Institution": nm, "Citations": ct, "Self-cite": "✓" if slf else ""}
                  for nm, ct, slf in citing["institutions"]]
        _crows = [{"Country": country_name(cc), "Citations": ct}
                  for cc, ct in citing["countries"]]
        _d = citing.get("_dbg", {})
        _note = mo.md(
            "*Captured **{}** citing records across **{}** papers "
            "(expected ~{} citations) · {} · {} API calls.*<br>"
            "*“2-yr mean citedness” is OpenAlex's open impact measure, calculated "
            "like an impact factor; the Journal Impact Factor itself is proprietary "
            "to Clarivate and has no open API.*".format(
                _d.get("captured"), _d.get("papers"), _d.get("expected"),
                _d.get("mode"), _d.get("calls")))
        _view = mo.vstack([_note, mo.ui.tabs({
            "Journals": mo.ui.table(_jrows, selection=None, pagination=True, page_size=25),
            "Institutions": mo.ui.table(_irows, selection=None, pagination=True, page_size=25),
            "Countries": mo.ui.table(_crows, selection=None, pagination=True, page_size=25),
        })])
    else:
        _d = citing.get("_dbg", {})
        _view = mo.md(
            "*No citing data came back.*\n\nDiagnostic — **{}** papers · "
            "**{}** calls (**{}** returned data) · captured **{}** of ~{} "
            "expected · mode: {}.\n\n`{}`".format(
                _d.get("papers"), _d.get("calls"), _d.get("ok"), _d.get("captured"),
                _d.get("expected"), _d.get("mode"), _d.get("sample")))

    mo.vstack([mo.md("### Who's citing your portfolio"), _view])
    return


if __name__ == "__main__":
    app.run()
